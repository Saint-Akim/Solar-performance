import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import requests
import io
import openpyxl
from datetime import datetime, timedelta
import concurrent.futures

st.set_page_config(page_title="Durr Bottling Electrical Analysis", page_icon="⚡", layout="wide", initial_sidebar_state="expanded")

def apply_professional_theme(dark_mode=True):
    if dark_mode:
        bg_main = "#1e1e2f"
        bg_card = "#2a2c41"
        text_color = "#ffffff"
        accent = "#4fd1c5"
        secondary_text = "#a0a0b0"
        grid_color = "#444444"
        line_color = "#f97316"
    else:
        bg_main = "#f5f5f8"
        bg_card = "#ffffff"
        text_color = "#000000"
        accent = "#4fd1c5"
        secondary_text = "#555555"
        grid_color = "#e0e0e0"
        line_color = "#f97316"
    css = f"""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@400;500;700&display=swap');
        .stApp {{ background-color: {bg_main}; color: {text_color}; font-family: 'Roboto', sans-serif; }}
        .fuel-card {{ background-color: {bg_card}; color: {text_color}; border-radius: 12px; padding: 24px; box-shadow: 0 4px 16px rgba(0,0,0,0.08); margin: 16px 0; }}
        .stMetric {{ background-color: {bg_card}; color: {text_color}; border-radius: 12px; padding: 25px 20px; box-shadow: 0 8px 16px rgba(0,0,0,0.2); margin-bottom: 15px; }}
        .stPlotlyChart iframe {{ border-radius: 12px !important; background-color: {bg_card} !important; }}
        .stSidebar {{ background-color: {bg_card}; color: {text_color}; padding: 20px; }}
        .stButton>button {{ background-color: {accent}; color: {text_color}; border-radius: 8px; font-weight: bold; padding: 10px 15px; border: none; }}
        .stTabs [data-baseweb="tab"][aria-selected="true"] {{ color: {accent}; border-bottom: 4px solid {accent}; }}
        .secondary-text {{ color: {secondary_text}; font-size: 0.9em; }}
        #MainMenu, footer, header {{ visibility: hidden; }}
    </style>
    """
    st.markdown(css, unsafe_allow_html=True)

with st.sidebar:
    st.markdown("## ⚡ Durr Bottling")
    st.caption("Electrical & Energy Intelligence")
    st.markdown("---")
    st.subheader("Appearance")
    col1, col2 = st.columns([0.7, 0.3])
    with col1:
        st.caption("Dark Mode")
    with col2:
        dark_mode = st.toggle("", value=True, label_visibility="collapsed")
        if 'theme' not in st.session_state or dark_mode != (st.session_state.theme == 'dark'):
            st.session_state.theme = 'dark' if dark_mode else 'light'
            st.rerun()
    apply_professional_theme(dark_mode)
    st.markdown("---")
    st.subheader("Date Range")
    preset = st.selectbox("Quick Select", ["Custom", "Last 7 Days", "Last 30 Days", "Last 90 Days", "Year to Date"])
    today = datetime.today().date()
    if preset == "Last 7 Days":
        start_date = today - timedelta(days=6)
        end_date = today
    elif preset == "Last 30 Days":
        start_date = today - timedelta(days=29)
        end_date = today
    elif preset == "Last 90 Days":
        start_date = today - timedelta(days=89)
        end_date = today
    elif preset == "Year to Date":
        start_date = datetime(today.year, 1, 1).date()
        end_date = today
    else:
        start_date = st.date_input("From", datetime(2025, 1, 1))
        end_date = st.date_input("To", today)
    if end_date < start_date:
        st.error("End date must be after start date")
        st.stop()
    st.markdown("---")
    st.subheader("Generator")
    fuel_region = st.selectbox("Fuel Region", ["Coast", "Reef"], help="Paarl = Coastal pricing")
    st.markdown("---")

SOLAR_URLS = [
    "https://raw.githubusercontent.com/Saint-Akim/Solar-performance/main/Solar_Goodwe%26Fronius-Jan.csv",
    "https://raw.githubusercontent.com/Saint-Akim/Solar-performance/main/Sloar_Goodwe%26Fronius_Feb.csv",
    "https://raw.githubusercontent.com/Saint-Akim/Solar-performance/main/Sloar_Goddwe%26Fronius_March.csv",
    "https://raw.githubusercontent.com/Saint-Akim/Solar-performance/main/Solar_goodwe%26Fronius_April.csv",
    "https://raw.githubusercontent.com/Saint-Akim/Solar-performance/main/Solar_goodwe%26Fronius_may.csv"
]
GEN_URL = "https://raw.githubusercontent.com/Saint-Akim/Solar-performance/main/gen%20(2).csv"
FACTORY_URL = "https://raw.githubusercontent.com/Saint-Akim/Solar-performance/main/FACTORY%20ELEC.csv"
KEHUA_URL = "https://raw.githubusercontent.com/Saint-Akim/Solar-performance/main/KEHUA%20INTERNAL.csv"
BILLING_URL = "https://raw.githubusercontent.com/Saint-Akim/Solar-performance/main/September%202025.xlsx"
HISTORY_URL = "https://raw.githubusercontent.com/Saint-Akim/Solar-performance/main/history.csv"
FUEL_PURCHASE_URL = "https://raw.githubusercontent.com/Saint-Akim/Solar-performance/main/Durr%20bottling%20Generator%20filling.xlsx"
FUEL_LEVEL_URL = "https://raw.githubusercontent.com/Saint-Akim/Solar-performance/main/history%20(5).csv"

def fetch_clean_data(args):
    url, is_generator = args
    try:
        df = pd.read_csv(url)
        if not is_generator and {'last_changed', 'state', 'entity_id'}.issubset(df.columns):
            df['last_changed'] = pd.to_datetime(df['last_changed'], utc=True).dt.tz_convert('Africa/Johannesburg').dt.tz_localize(None)
            df['state'] = pd.to_numeric(df['state'], errors='coerce').abs()
            df['entity_id'] = df['entity_id'].str.lower().str.strip()
            return df.pivot_table(index='last_changed', columns='entity_id', values='state', aggfunc='mean').reset_index()
        return df
    except Exception as e:
        st.warning(f"Error loading {url}: {e}")
        return pd.DataFrame()

@st.cache_data(show_spinner="Loading data from GitHub...")
def load_data_engine():
    urls = SOLAR_URLS + [GEN_URL, FACTORY_URL, KEHUA_URL, HISTORY_URL, FUEL_LEVEL_URL]
    flags = [False] * len(SOLAR_URLS) + [True, False, False, False, False]
    with concurrent.futures.ThreadPoolExecutor() as executor:
        results = list(executor.map(fetch_clean_data, zip(urls, flags)))
    solar_df = pd.concat([r for r in results[:len(SOLAR_URLS)] if not r.empty], ignore_index=True) if any(not r.empty for r in results[:len(SOLAR_URLS)]) else pd.DataFrame()
    gen_df = results[len(SOLAR_URLS)]
    factory_df = results[len(SOLAR_URLS)+1]
    kehua_df = results[len(SOLAR_URLS)+2]
    history_df = results[len(SOLAR_URLS)+3]
    fuel_level_df = results[-1]
    return solar_df, gen_df, factory_df, kehua_df, history_df, fuel_level_df

solar_df, gen_df, factory_df, kehua_df, history_df, fuel_level_df = load_data_engine()

# ------------------ LOAD FUEL PURCHASES (SAFE DATE HANDLING) ------------------
def load_fuel_purchases():
    try:
        resp = requests.get(FUEL_PURCHASE_URL)
        resp.raise_for_status()
        buffer = io.BytesIO(resp.content)
        df = pd.read_excel(buffer, sheet_name=0, header=0)
    except Exception as e:
        st.warning(f"Error loading fuel purchases: {e}")
        return pd.DataFrame()
    df.columns = df.columns.str.lower().str.replace(' ', '_')
    def excel_to_date(val):
        if pd.isna(val):
            return pd.NaT
        if isinstance(val, (int, float)):
            return datetime(1899, 12, 30) + timedelta(days=val)
        return pd.to_datetime(val, errors='coerce')
    df['date'] = df['date'].apply(excel_to_date)
    df = df.dropna(subset=['date'])
    df.rename(columns={'amount(liters)': 'liters', 'price_per_litre': 'price_per_l', 'cost(rands)': 'cost_r'}, inplace=True)
    return df

fuel_purchases = load_fuel_purchases()
filtered_fuel = fuel_purchases[(fuel_purchases['date'].dt.date >= start_date) & (fuel_purchases['date'].dt.date <= end_date)] if not fuel_purchases.empty else pd.DataFrame()

# ------------------ DYNAMIC MONTHLY PRICES FROM FUEL PURCHASES ------------------
def get_monthly_prices(fuel_purchases_df):
    if fuel_purchases_df.empty:
        # Fallback to Coast prices (Paarl = Coast)
        return pd.DataFrame({
            'month': pd.to_datetime(['2025-01-01', '2025-02-01', '2025-03-01', '2025-04-01', '2025-05-01', '2025-06-01',
                                     '2025-07-01', '2025-08-01', '2025-09-01', '2025-10-01', '2025-11-01', '2025-12-01']),
            'price': [20.28, 21.62, 21.55, 20.79, 20.57, 17.81, 18.53, 19.20, 19.80, 19.50, 19.00, 22.52]
        })
   
    fuel_purchases_df = fuel_purchases_df.copy()
    fuel_purchases_df['month'] = fuel_purchases_df['date'].dt.to_period('M').dt.to_timestamp()
    monthly = fuel_purchases_df.groupby('month').apply(
        lambda g: pd.Series({
            'total_liters': g['liters'].sum(),
            'total_cost': g['cost_r'].sum()
        })
    ).reset_index()
    monthly['price'] = monthly['total_cost'] / monthly['total_liters']
    monthly = monthly[['month', 'price']]
   
    all_months = pd.date_range('2025-01-01', '2025-12-01', freq='MS').to_series().dt.to_period('M').dt.to_timestamp()
    monthly = pd.merge(pd.DataFrame({'month': all_months}), monthly, on='month', how='left')
    monthly['price'] = monthly['price'].ffill().bfill()
   
    return monthly

historical_prices = get_monthly_prices(fuel_purchases)
current_price = historical_prices['price'].iloc[-1] if not historical_prices.empty else 22.52

# ------------------ GENERATOR DATA PROCESSING ------------------
@st.cache_data(show_spinner=False)
def process_generator_data(gen_df: pd.DataFrame, fuel_level_df: pd.DataFrame, history_df: pd.DataFrame, fuel_purchases: pd.DataFrame):
    # Process gen (2).csv - Cumulative fuel consumed
    if not gen_df.empty:
        gen_df = gen_df.copy()
        gen_df['last_changed'] = pd.to_datetime(gen_df['last_changed'])
        gen_df = gen_df.sort_values('last_changed')
        gen_df['fuel_consumed'] = gen_df['state'].diff().clip(lower=0).fillna(0)
    else:
        gen_df = pd.DataFrame()

    # Process history (5).csv - Fuel level start (tank level)
    if not fuel_level_df.empty:
        fuel_level_df = fuel_level_df.copy()
        fuel_level_df['last_changed'] = pd.to_datetime(fuel_level_df['last_changed'])
        fuel_level_df = fuel_level_df.sort_values('last_changed')
        fuel_level_df['fuel_level_change'] = -fuel_level_df['state'].diff().clip(lower=0).fillna(0)  # Negative diff for consumption
        fuel_level_df['refill'] = fuel_level_df['state'].diff() > 0
    else:
        fuel_level_df = pd.DataFrame()

    # Process history.csv - Runtime duration
    if not history_df.empty:
        history_df = history_df.copy()
        history_df['last_changed'] = pd.to_datetime(history_df['last_changed'])
        history_df = history_df.sort_values('last_changed')
        history_df['runtime_change'] = history_df['state'].diff().clip(lower=0).fillna(0)
    else:
        history_df = pd.DataFrame()

    # Merge all generator data on time
    all_gen = pd.concat([gen_df, fuel_level_df, history_df], ignore_index=True, sort=False)
    all_gen = all_gen.sort_values('last_changed').drop_duplicates(subset=['last_changed'], keep='last')

    # Fill missing columns with 0
    all_gen['fuel_consumed'] = all_gen.get('fuel_consumed', 0).fillna(0)
    all_gen['fuel_level_change'] = all_gen.get('fuel_level_change', 0).fillna(0)
    all_gen['runtime_change'] = all_gen.get('runtime_change', 0).fillna(0)

    # Daily aggregation
    daily = all_gen.resample('D', on='last_changed').agg({
        'fuel_consumed': 'sum',
        'fuel_level_change': 'sum',
        'runtime_change': 'sum',
        'refill': 'sum'
    }).reset_index()

    # Fuel efficiency calculation: use max of fuel_consumed or fuel_level_change for accuracy
    daily['fuel_used_l'] = daily[['fuel_consumed', 'fuel_level_change']].max(axis=1)
    daily['runtime_h'] = daily['runtime_change']
    daily['liters_per_hour'] = daily.apply(lambda r: r['fuel_used_l'] / r['runtime_h'] if r['runtime_h'] > 0 else 0, axis=1)
    daily['month'] = daily['last_changed'].dt.to_period('M').dt.to_timestamp()
    daily = daily.merge(historical_prices[['month', 'price']], on='month', how='left')
    daily['price_final'] = daily['price'].fillna(current_price)
    daily['daily_cost_r'] = daily['fuel_used_l'] * daily['price_final']
    daily.drop(columns=['month', 'price'], inplace=True)

    totals = {
        "total_cost_r": daily['daily_cost_r'].sum(),
        "total_fuel_l": daily['fuel_used_l'].sum(),
        "total_runtime_h": daily['runtime_h'].sum(),
        "avg_liters_per_hour": daily['liters_per_hour'][daily['liters_per_hour'] > 0].mean() if (daily['liters_per_hour'] > 0).any() else 0,
        "total_refills": daily['refill'].sum()
    }

    return daily, totals

daily_gen, totals = process_generator_data(gen_df, fuel_level_df, history_df, fuel_purchases)
if not daily_gen.empty:
    filtered_gen = daily_gen[
        (daily_gen['last_changed'].dt.date >= start_date) &
        (daily_gen['last_changed'].dt.date <= end_date)
    ].copy()
    filtered_totals = {
        "total_cost_r": filtered_gen['daily_cost_r'].sum(),
        "total_fuel_l": filtered_gen['fuel_used_l'].sum(),
        "total_runtime_h": filtered_gen['runtime_h'].sum(),
        "avg_liters_per_hour": filtered_gen['liters_per_hour'][filtered_gen['liters_per_hour'] > 0].mean() if (filtered_gen['liters_per_hour'] > 0).any() else 0,
        "total_refills": filtered_gen['refill'].sum()
    }
else:
    filtered_gen = pd.DataFrame()
    filtered_totals = {}

# ------------------ TABS ------------------
tab1, tab2, tab3, tab4 = st.tabs(["🔌 Generator Cost & Fuel", "☀️ Solar Performance", "🏭 Factory Load", "📄 Billing Editor"])

# ==================== GENERATOR ANALYSIS TAB ====================
with tab1:
    st.markdown("## 🔌 Generator Performance Overview")
    st.caption("Fuel usage, runtime and estimated operating cost using actual purchase prices")
    st.markdown("<div class='fuel-card'>", unsafe_allow_html=True)
    st.markdown(f"### Current Diesel Price: **R {current_price:.2f}/L** (from latest purchase)")
    if filtered_totals:
        cols = st.columns(6)
        cols[0].metric("Estimated Total Cost", f"R {filtered_totals['total_cost_r']:,.0f}")
        cols[1].metric("Estimated Fuel Used", f"{filtered_totals['total_fuel_l']:.1f} L")
        cols[2].metric("Runtime", f"{filtered_totals['total_runtime_h']:.1f} h")
        cols[3].metric("Avg Liters per Hour", f"{filtered_totals['avg_liters_per_hour']:.2f} L/h")
        cols[4].metric("Daily Avg Cost", f"R {filtered_gen['daily_cost_r'].mean():,.0f}")
        cols[5].metric("Refills", f"{filtered_totals['total_refills']}")
        fig = px.bar(filtered_gen, x='last_changed', y='fuel_used_l', text='fuel_used_l', color_discrete_sequence=["#4fd1c5"])
        fig.add_scatter(x=filtered_gen['last_changed'], y=filtered_gen['liters_per_hour'] * 20, mode="lines+markers", name="L/h (scaled x20)", line=dict(color="#f97316", width=3))
        fig.update_layout(
            height=520,
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(rangeslider=dict(visible=True), showgrid=False),
            yaxis=dict(title="Fuel Used (L)", gridcolor="#444444" if dark_mode else "#e0e0e0"),
            legend=dict(orientation="h")
        )
        st.plotly_chart(fig, use_container_width=True)
        lph = filtered_totals['avg_liters_per_hour']
        if lph > 0:
            if lph < 3.5:
                st.success(f"Average Consumption: {lph:.2f} L/h (Excellent)")
            elif lph < 4.0:
                st.success(f"Average Consumption: {lph:.2f} L/h (Good)")
            elif lph < 4.5:
                st.warning(f"Average Consumption: {lph:.2f} L/h (Average)")
            else:
                st.error(f"Average Consumption: {lph:.2f} L/h (High)")
        st.caption("⚠️ Estimated cost based on actual monthly purchase prices. Variance from actual spend may be due to timing.")
    st.markdown("### Actual Fuel Purchases")
    if not filtered_fuel.empty:
        purchase_liters = filtered_fuel['liters'].sum()
        purchase_cost = filtered_fuel['cost_r'].sum()
        avg_price = purchase_cost / purchase_liters if purchase_liters > 0 else 0
        cols = st.columns(4)
        cols[0].metric("Actual Liters Purchased", f"{purchase_liters:.1f} L")
        cols[1].metric("Actual Total Cost", f"R {purchase_cost:,.0f}")
        cols[2].metric("Avg Price per Liter", f"R {avg_price:.2f}")
        cols[3].metric("Purchase Events", f"{len(filtered_fuel)}")
        fig_p = px.bar(filtered_fuel, x='date', y='liters', color='price_per_l', text='liters', color_continuous_scale="viridis")
        fig_p.update_layout(title="Fuel Purchase Events", xaxis_title="Date", yaxis_title="Liters", coloraxis_showscale=False)
        st.plotly_chart(fig_p, use_container_width=True)
        variance = filtered_totals.get('total_cost_r', 0) - purchase_cost
        if abs(variance) < 100:
            st.success("Estimated and Actual costs closely match!")
        elif variance > 0:
            st.warning(f"Estimated exceeds Actual by R {variance:,.0f}")
        else:
            st.info(f"Actual exceeds Estimated by R {abs(variance):,.0f}")
        with st.expander("Purchase Details"):
            st.dataframe(filtered_fuel.style.format({'liters': '{:.1f}', 'price_per_l': 'R {:.2f}', 'cost_r': 'R {:.2f}'}))
    else:
        st.info("No fuel purchase data in selected period.")
    st.markdown("</div>", unsafe_allow_html=True)

# ==================== SOLAR PERFORMANCE TAB ====================
with tab2:
    st.markdown("<div class='fuel-card'>", unsafe_allow_html=True)
    if not solar_df.empty:
        solar_df['last_changed'] = pd.to_datetime(solar_df['last_changed'])
        filtered_solar = solar_df[(solar_df['last_changed'].dt.date >= start_date) & (solar_df['last_changed'].dt.date <= end_date)]
        if not filtered_solar.empty:
            fig = go.Figure(go.Scatter(
                x=filtered_solar['last_changed'], y=filtered_solar.get('sensor.goodwe_grid_power', 0)/1000 + filtered_solar.get('sensor.fronius_grid_power', 0)/1000,
                mode='lines', name='Total Solar Output',
                line=dict(color="#4fd1c5", width=3),
                fill='tozeroy'
            ))
            fig.update_layout(title="Solar Power Output (kW)", height=500, xaxis=dict(rangeslider=dict(visible=True)), plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig, use_container_width=True)
            peak = (filtered_solar.get('sensor.goodwe_grid_power', 0) + filtered_solar.get('sensor.fronius_grid_power', 0)).max() / 1000
            st.success(f"Peak Solar Output: {peak:.1f} kW")
            filtered_solar['hour'] = filtered_solar['last_changed'].dt.hour
            heatmap_data = filtered_solar.groupby('hour')[['sensor.goodwe_grid_power', 'sensor.fronius_grid_power']].mean().sum(axis=1)/1000
            heatmap_data = heatmap_data.reset_index(name='total_solar')
            fig_heat = px.bar(heatmap_data, x='hour', y='total_solar', title="Average Solar Output by Hour of Day", color_discrete_sequence=["#4fd1c5"])
            st.plotly_chart(fig_heat, use_container_width=True)
    else:
        st.info("No solar data available.")
    st.markdown("</div>", unsafe_allow_html=True)

# ==================== FACTORY LOAD TAB ====================
with tab3:
    st.markdown("<div class='fuel-card'>", unsafe_allow_html=True)
    if not factory_df.empty:
        factory_filtered = factory_df.copy()
        if 'last_changed' in factory_filtered.columns:
            factory_filtered['last_changed'] = pd.to_datetime(factory_filtered['last_changed'])
            factory_filtered = factory_filtered.sort_values('last_changed')
            if 'sensor.bottling_factory_monthkwhtotal' in factory_filtered.columns:
                factory_filtered['daily_kwh'] = factory_filtered['sensor.bottling_factory_monthkwhtotal'].diff().clip(lower=0).fillna(0)
                factory_filtered = factory_filtered[
                    (factory_filtered['last_changed'].dt.date >= start_date) &
                    (factory_filtered['last_changed'].dt.date <= end_date)
                ]
                fig = go.Figure(go.Bar(
                    x=factory_filtered['last_changed'], y=factory_filtered['daily_kwh'],
                    marker_color="#3498DB"
                ))
                fig.update_layout(title="Daily Factory Energy Consumption (kWh)", height=500, xaxis=dict(rangeslider=dict(visible=True)), plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig, use_container_width=True)
                st.metric("Total Factory kWh", f"{factory_filtered['daily_kwh'].sum():,.0f} kWh")
            else:
                st.info("Factory cumulative kWh sensor not found.")
        else:
            st.info("No timestamp column in factory data.")
    else:
        st.info("No factory consumption data available.")
    st.markdown("</div>", unsafe_allow_html=True)

# ==================== BILLING EDITOR TAB ====================
with tab4:
    st.markdown("<div class='fuel-card'>", unsafe_allow_html=True)
    st.subheader("September 2025 Invoice Editor")
    try:
        resp = requests.get(BILLING_URL)
        resp.raise_for_status()
        buffer = io.BytesIO(resp.content)
        wb = openpyxl.load_workbook(buffer, data_only=False)
        ws = wb.active
        col1, col2 = st.columns(2)
        with col1:
            from_val = ws['B2'].value or "30/09/25"
            if isinstance(from_val, datetime):
                from_date = from_val.date()
            else:
                from_date = datetime.strptime(str(from_val).strip(), "%d/%m/%y").date()
            from_date = st.date_input("Period From (B2)", value=from_date)
            freedom_units = float(ws['C7'].value or 0)
            freedom_units = st.number_input("Freedom Village Units (C7)", value=freedom_units)
        with col2:
            to_val = ws['B3'].value or "31/10/25"
            if isinstance(to_val, datetime):
                to_date = to_val.date()
            else:
                to_date = datetime.strptime(str(to_val).strip(), "%d/%m/%y").date()
            to_date = st.date_input("Period To (B3)", value=to_date)
            boerdery_units = float(ws['C9'].value or 0)
            boerdery_units = st.number_input("Boerdery Units (C9)", value=boerdery_units)
        if st.button("Generate Updated Invoice", type="primary"):
            ws['B2'].value = from_date.strftime("%d/%m/%y")
            ws['B3'].value = to_date.strftime("%d/%m/%y")
            ws['C7'].value = freedom_units
            ws['C9'].value = boerdery_units
            new_buffer = io.BytesIO()
            wb.save(new_buffer)
            new_buffer.seek(0)
            st.download_button(
                label="Download Updated Invoice",
                data=new_buffer,
                file_name=f"Invoice_{from_date.strftime('%b%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception as e:
        st.error(f"Failed to load billing template: {e}")
    st.markdown("</div>", unsafe_allow_html=True)

# ------------------ FOOTER ------------------
st.markdown("---")
st.markdown("<p style='text-align:center; font-weight:bold;'>built by Electrical@durrbottling.com</p>", unsafe_allow_html=True)
